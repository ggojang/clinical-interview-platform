#!/usr/bin/env python3
"""Build and manage the isolated test health-screening package catalog.

This tool intentionally writes outside clinical Knowledge/Fact packages.  The
generated catalog is public listing data for Custom GPT Action testing, not an
independent clinical recommendation authority.
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
import hashlib
import json
from pathlib import Path
import re
import shutil
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = (
    ROOT / "docs" / "gpt" / "test-catalogs" / "health-screening-packages"
)
CATALOG_ID = "test.kr-health-screening-center-packages"
REGION_PAGE_SIZE = 50
REGION_IDS = {
    "서울": "seoul",
    "경기": "gyeonggi",
    "인천": "incheon",
    "부산": "busan",
    "대구": "daegu",
    "광주": "gwangju",
    "대전": "daejeon",
    "울산": "ulsan",
    "세종": "sejong",
    "강원": "gangwon",
    "충북": "chungbuk",
    "충남": "chungnam",
    "전북": "jeonbuk",
    "전남": "jeonnam",
    "경북": "gyeongbuk",
    "경남": "gyeongnam",
    "제주": "jeju",
}
EXPECTED_COLUMNS = [
    "지역",
    "분류",
    "기관명",
    "패키지명",
    "대상",
    "주요 검사항목",
    "가격(원)",
    "소요시간",
    "비고",
    "수집상태",
    "출처 URL",
]

TAG_PATTERNS = {
    "brain-mri-mra": ("뇌mri", "뇌 mri", "뇌mra", "뇌 mra"),
    "pet-ct": ("pet-ct", "pet ct", "펫ct", "펫 ct"),
    "low-dose-chest-ct": ("저선량 흉부", "저선량폐", "저선량 폐"),
    "ct": ("ct", "컴퓨터단층"),
    "mri-mra": ("mri", "mra", "자기공명"),
    "gastroscopy": ("위내시경", "위 내시경"),
    "colonoscopy": ("대장내시경", "대장 내시경"),
    "abdominal-ultrasound": ("복부초음파", "복부 초음파"),
    "thyroid-ultrasound": ("갑상선초음파", "갑상선 초음파"),
    "carotid-ultrasound": ("경동맥초음파", "경동맥 초음파"),
    "breast-screening": ("유방", "맘모"),
    "cervical-screening": ("자궁경부", "부인과"),
    "prostate-screening": ("전립선", "psa"),
    "cardiovascular": ("심혈관", "관상동맥", "심장"),
    "bone-density": ("골밀도",),
}


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace(
        "+00:00", "Z"
    )


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def load_registry(output: Path) -> dict[str, Any]:
    path = output / "registry.json"
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {
        "catalog_id": CATALOG_ID,
        "title_ko": "건강검진센터 추가 패키지 테스트 카탈로그",
        "lifecycle_status": "test",
        "review_status": "unreviewed",
        "clinical_use_status": "limited",
        "contains_patient_responses": False,
        "test_only": True,
        "current_version": None,
        "versions": [],
        "privacy_boundary": {
            "accepted_action_parameters": [
                "catalogVersion",
                "regionId",
                "packageId",
            ],
            "forbidden_action_payloads": [
                "patient_answers",
                "free_text_health_information",
                "age",
                "sex_or_gender",
                "diagnoses",
                "medications",
                "budget",
                "identifiers",
            ],
        },
        "removal": {
            "command": "python3 tools/test_catalogs/build_health_screening_package_catalog.py remove --version {version}",
            "action_schema_paths_are_isolated": True,
        },
    }


def source_digest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def price_projection(raw: str | None) -> dict[str, Any]:
    if not raw:
        return {
            "raw": None,
            "kind": "unavailable",
            "minimum_krw": None,
            "maximum_krw": None,
            "listed_variants": [],
        }
    numbers = [
        int(token.replace(",", ""))
        for token in re.findall(r"(?<!\d)\d[\d,]*", raw)
    ]
    if not numbers:
        kind = "unavailable"
    elif len(numbers) == 1:
        kind = "single"
    elif re.search(r"[~～–—]", raw):
        kind = "range"
    elif "/" in raw:
        kind = "listed_variants"
    else:
        kind = "multiple_or_range_unspecified"
    variants = []
    if kind == "listed_variants":
        for index, segment in enumerate(raw.split("/"), start=1):
            match = re.search(r"\d[\d,]*", segment)
            if not match:
                continue
            label = (segment[: match.start()] + segment[match.end() :]).strip(" :원()")
            variants.append(
                {
                    "sequence": index,
                    "label_raw": label or None,
                    "amount_krw": int(match.group(0).replace(",", "")),
                }
            )
    return {
        "raw": raw,
        "kind": kind,
        "minimum_krw": min(numbers) if numbers else None,
        "maximum_krw": max(numbers) if numbers else None,
        "listed_variants": variants,
    }


def lexical_tags(*values: str | None) -> list[str]:
    text = " ".join(value or "" for value in values).lower()
    return sorted(
        tag
        for tag, patterns in TAG_PATTERNS.items()
        if any(pattern in text for pattern in patterns)
    )


def package_identifier(
    region: str, category: str | None, institution: str, package_name: str
) -> str:
    semantic_key = "|".join(
        [region.strip(), (category or "").strip(), institution.strip(), package_name.strip()]
    )
    return "pkg-" + hashlib.sha256(semantic_key.encode("utf-8")).hexdigest()[:16]


def read_rows(input_path: Path) -> tuple[list[dict[str, Any]], str, int]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if "전체패키지" not in workbook.sheetnames:
        raise ValueError("workbook must contain the 전체패키지 sheet")
    sheet = workbook["전체패키지"]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    if headers != EXPECTED_COLUMNS:
        raise ValueError(f"unexpected columns: {headers!r}")
    values = []
    for row_number, row in enumerate(rows, start=2):
        if len(row) != len(EXPECTED_COLUMNS):
            raise ValueError(
                f"unexpected column count at source row {row_number}: {len(row)}"
            )
        item = dict(zip(EXPECTED_COLUMNS, row))
        package_name = clean(item["패키지명"])
        institution = clean(item["기관명"])
        region = clean(item["지역"])
        if not package_name or not institution or not region:
            continue
        values.append(
            {
                "source_row": row_number,
                "region": region,
                "category": clean(item["분류"]),
                "institution": institution,
                "package_name": package_name,
                "target_text": clean(item["대상"]),
                "items_text": clean(item["주요 검사항목"]),
                "price_raw": clean(item["가격(원)"]),
                "duration_text": clean(item["소요시간"]),
                "notes": clean(item["비고"]),
                "source_status": clean(item["수집상태"]),
                "source_url": clean(item["출처 URL"]),
            }
        )
    return values, sheet.title, max(sheet.max_row - 1, 0)


def build_catalog(
    input_path: Path, output: Path, version: str, activate: bool
) -> dict[str, Any]:
    if not re.fullmatch(r"[0-9A-Za-z][0-9A-Za-z._-]{0,63}", version):
        raise ValueError("version must be a path-safe 1..64 character token")
    version_dir = output / "versions" / version
    if version_dir.exists():
        raise FileExistsError(
            f"version already exists: {version}; remove it explicitly before rebuilding"
        )
    rows, sheet_name, source_row_count = read_rows(input_path)
    generated_at = now_iso()
    digest = source_digest(input_path)
    packages: list[dict[str, Any]] = []
    regions: dict[str, list[dict[str, Any]]] = {}
    institutions: set[str] = set()
    grouped_rows: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        region_name = row["region"]
        if region_name not in REGION_IDS:
            raise ValueError(f"unsupported region label: {region_name}")
        package_id = package_identifier(
            region_name,
            row["category"],
            row["institution"],
            row["package_name"],
        )
        grouped_rows.setdefault(package_id, []).append(row)

    for package_id, variants_source in sorted(grouped_rows.items()):
        row = variants_source[0]
        region_name = row["region"]
        institutions.add(row["institution"])
        variants = []
        tag_set: set[str] = set()
        minimums = []
        maximums = []
        target_texts = []
        listing_statuses = []
        for sequence, source_row in enumerate(variants_source, start=1):
            price = price_projection(source_row["price_raw"])
            if price["minimum_krw"] is not None:
                minimums.append(price["minimum_krw"])
            if price["maximum_krw"] is not None:
                maximums.append(price["maximum_krw"])
            tags = lexical_tags(
                source_row["category"],
                source_row["package_name"],
                source_row["target_text"],
                source_row["items_text"],
            )
            tag_set.update(tags)
            if source_row["target_text"] not in target_texts:
                target_texts.append(source_row["target_text"])
            if source_row["source_status"] not in listing_statuses:
                listing_statuses.append(source_row["source_status"])
            variants.append(
                {
                    "variant_sequence": sequence,
                    "target_text": source_row["target_text"],
                    "items_text": source_row["items_text"],
                    "price": price,
                    "duration_text": source_row["duration_text"],
                    "notes": source_row["notes"],
                    "lexical_tags": tags,
                    "listing_status": source_row["source_status"],
                    "source": {
                        "sheet": sheet_name,
                        "row": source_row["source_row"],
                        "url": source_row["source_url"],
                    },
                }
            )
        aggregate_price = {
            "kind": "variant_collection" if len(variants) > 1 else variants[0]["price"]["kind"],
            "minimum_krw": min(minimums) if minimums else None,
            "maximum_krw": max(maximums) if maximums else None,
            "variant_count": len(variants),
        }
        tags = sorted(tag_set)
        detail = {
            "resource_type": "TestHealthScreeningPackage",
            "catalog_id": CATALOG_ID,
            "catalog_version": version,
            "package_id": package_id,
            "region": {
                "id": REGION_IDS[region_name],
                "display_ko": region_name,
            },
            "category": row["category"],
            "institution": row["institution"],
            "package_name": row["package_name"],
            "target_texts": target_texts,
            "price_summary": aggregate_price,
            "variants": variants,
            "lexical_tags": tags,
            "listing_statuses": listing_statuses,
            "source": {
                "workbook_sha256": digest,
                "sheet": sheet_name,
                "rows": [item["source_row"] for item in variants_source],
                "urls": sorted(
                    {item["source_url"] for item in variants_source if item["source_url"]}
                ),
            },
            "use_boundary": {
                "test_only": True,
                "unreviewed": True,
                "candidate_comparison_only": True,
                "not_clinical_advice": True,
                "confirm_current_items_price_and_eligibility_with_institution": True,
            },
        }
        packages.append(detail)
        summary = {
            "package_id": package_id,
            "institution": row["institution"],
            "package_name": row["package_name"],
            "category": row["category"],
            "target_texts": target_texts,
            "price_summary": aggregate_price,
            "variant_count": len(variants),
            "lexical_tags": tags,
            "listing_statuses": listing_statuses,
            "detail_path": (
                f"/gpt/test-catalogs/health-screening-packages/versions/"
                f"{version}/packages/{package_id}.json"
            ),
        }
        regions.setdefault(REGION_IDS[region_name], []).append(summary)

    for package in packages:
        write_json(
            version_dir / "packages" / f"{package['package_id']}.json",
            package,
        )
    region_entries = []
    for display_ko, region_id in REGION_IDS.items():
        region_packages = sorted(
            regions.get(region_id, []),
            key=lambda item: (
                item["price_summary"]["minimum_krw"] is None,
                item["price_summary"]["minimum_krw"] or 0,
                item["institution"],
                item["package_name"],
            ),
        )
        page_entries = []
        for offset in range(0, len(region_packages), REGION_PAGE_SIZE):
            page_number = offset // REGION_PAGE_SIZE + 1
            page_packages = region_packages[offset : offset + REGION_PAGE_SIZE]
            page_document = {
                "resource_type": "TestHealthScreeningPackageRegionPage",
                "catalog_id": CATALOG_ID,
                "catalog_version": version,
                "region": {"id": region_id, "display_ko": display_ko},
                "page": page_number,
                "page_size": REGION_PAGE_SIZE,
                "package_count": len(page_packages),
                "packages": page_packages,
                "contains_patient_responses": False,
            }
            page_path = (
                f"/gpt/test-catalogs/health-screening-packages/versions/"
                f"{version}/regions/{region_id}/pages/{page_number}.json"
            )
            write_json(
                version_dir
                / "regions"
                / region_id
                / "pages"
                / f"{page_number}.json",
                page_document,
            )
            page_entries.append(
                {
                    "page": page_number,
                    "package_count": len(page_packages),
                    "path": page_path,
                }
            )
        region_document = {
            "resource_type": "TestHealthScreeningPackageRegionMetadata",
            "catalog_id": CATALOG_ID,
            "catalog_version": version,
            "region": {"id": region_id, "display_ko": display_ko},
            "package_count": len(region_packages),
            "page_size": REGION_PAGE_SIZE,
            "page_count": len(page_entries),
            "pages": page_entries,
            "sorting": "known minimum price ascending, then institution and package name",
            "contains_patient_responses": False,
        }
        write_json(version_dir / "regions" / region_id / "index.json", region_document)
        region_entries.append(
            {
                "id": region_id,
                "display_ko": display_ko,
                "package_count": len(region_packages),
                "page_count": len(page_entries),
                "path": (
                    f"/gpt/test-catalogs/health-screening-packages/versions/"
                    f"{version}/regions/{region_id}/index.json"
                ),
            }
        )

    metadata = {
        "resource_type": "TestHealthScreeningPackageCatalogMetadata",
        "catalog_id": CATALOG_ID,
        "catalog_version": version,
        "generated_at": generated_at,
        "lifecycle_status": "test",
        "review_status": "unreviewed",
        "clinical_use_status": "limited",
        "contains_patient_responses": False,
        "source": {
            "kind": "user_supplied_workbook_of_public_center_listings",
            "filename": input_path.name,
            "sheet": sheet_name,
            "sha256": digest,
        },
        "counts": {
            "packages": len(packages),
            "listing_variants": len(rows),
            "institutions": len(institutions),
            "regions": len(REGION_IDS),
            "source_rows": source_row_count,
            "source_rows_without_named_package": source_row_count - len(rows),
        },
        "regions": region_entries,
        "path_templates": {
            "region": (
                "/gpt/test-catalogs/health-screening-packages/versions/"
                "{catalogVersion}/regions/{regionId}/index.json"
            ),
            "region_page": (
                "/gpt/test-catalogs/health-screening-packages/versions/"
                "{catalogVersion}/regions/{regionId}/pages/{page}.json"
            ),
            "package": (
                "/gpt/test-catalogs/health-screening-packages/versions/"
                "{catalogVersion}/packages/{packageId}.json"
            ),
        },
        "interpretation_policy": {
            "source_text_is_preserved": True,
            "lexical_tags_are_search_aids_not_clinical_claims": True,
            "minimum_and_maximum_prices_are_parsed_from_price_raw": True,
            "listed_variants_are_not_assigned_to_sex_without_an_explicit_label": True,
            "institution_confirmation_required": True,
        },
    }
    write_json(version_dir / "metadata.json", metadata)

    registry = load_registry(output)
    registry["versions"] = [
        item for item in registry["versions"] if item["version"] != version
    ]
    registry["versions"].append(
        {
            "version": version,
            "status": "active" if activate else "inactive",
            "generated_at": generated_at,
            "source_sha256": digest,
            "package_count": len(packages),
            "region_count": len(REGION_IDS),
            "metadata_path": (
                f"/gpt/test-catalogs/health-screening-packages/versions/"
                f"{version}/metadata.json"
            ),
        }
    )
    if activate:
        registry["current_version"] = version
    for item in registry["versions"]:
        item["status"] = (
            "active" if item["version"] == registry["current_version"] else "inactive"
        )
    write_json(output / "registry.json", registry)
    return metadata


def activate_version(output: Path, version: str) -> None:
    registry = load_registry(output)
    versions = {item["version"] for item in registry["versions"]}
    if version not in versions or not (output / "versions" / version).exists():
        raise ValueError(f"catalog version does not exist: {version}")
    registry["current_version"] = version
    for item in registry["versions"]:
        item["status"] = "active" if item["version"] == version else "inactive"
    write_json(output / "registry.json", registry)


def remove_version(output: Path, version: str, force_current: bool) -> None:
    registry = load_registry(output)
    if registry.get("current_version") == version and not force_current:
        raise ValueError("cannot remove the active version without --force-current")
    shutil.rmtree(output / "versions" / version, ignore_errors=True)
    registry["versions"] = [
        item for item in registry["versions"] if item["version"] != version
    ]
    if registry.get("current_version") == version:
        registry["current_version"] = None
    write_json(output / "registry.json", registry)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    subparsers = parser.add_subparsers(dest="command", required=True)

    build_parser = subparsers.add_parser("build")
    build_parser.add_argument("--input", type=Path, required=True)
    build_parser.add_argument("--version", required=True)
    build_parser.add_argument("--activate", action="store_true")

    activate_parser = subparsers.add_parser("activate")
    activate_parser.add_argument("--version", required=True)

    remove_parser = subparsers.add_parser("remove")
    remove_parser.add_argument("--version", required=True)
    remove_parser.add_argument("--force-current", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    output = args.output.resolve()
    if args.command == "build":
        metadata = build_catalog(
            args.input.resolve(), output, args.version, args.activate
        )
        print(
            f"built {metadata['counts']['packages']} packages across "
            f"{metadata['counts']['regions']} regions: {args.version}"
        )
    elif args.command == "activate":
        activate_version(output, args.version)
        print(f"activated catalog version: {args.version}")
    else:
        remove_version(output, args.version, args.force_current)
        print(f"removed catalog version: {args.version}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
